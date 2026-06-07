"""
Excel/CSV to Notepad Converter
==============================
A robust script that converts Excel sheets or CSV files into readable .txt
files. Excel workbooks create one .txt file per sheet.

Dependencies: pandas, openpyxl
Install: pip install pandas openpyxl
"""

# Import argparse to handle command-line arguments and options
import argparse
# Import logging to track script execution and display status messages
import logging
# Import re for regular expression operations (filename sanitization)
import re
# Import textwrap to wrap long text lines to fit within specified widths
import textwrap
# Import Path from pathlib for cross-platform file path operations
from pathlib import Path
# Import type hints for better code readability and IDE support
from typing import Dict, List, Tuple

# Import pandas for data manipulation and Excel/CSV file reading
import pandas as pd
# Import FPDF for generating PDF output files
from fpdf import FPDF
# Import load_workbook from openpyxl to read Excel workbooks
from openpyxl import load_workbook
# Import get_column_letter to convert column numbers to Excel letters (e.g., 1 -> A)
from openpyxl.utils import get_column_letter


# Configure basic logging with INFO level and a simple message format
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
# Create a logger instance named after this module for consistent logging
logger = logging.getLogger(__name__)

# Compile a regex pattern to match characters invalid in filenames (cross-platform)
INVALID_FILENAME_CHARS = re.compile(r'[\\/*?:"<>|]')
# Define the minimum allowed width for formatted columns in text output
MIN_COLUMN_WIDTH = 2
# Define the default maximum width for columns before text wrapping occurs
DEFAULT_MAX_COLUMN_WIDTH = 60
# Define a set of file extensions supported for Excel files
SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
# Combine Excel extensions with CSV to create the full set of supported input types
SUPPORTED_INPUT_EXTENSIONS = SUPPORTED_EXCEL_EXTENSIONS | {".csv"}


def sanitize_filename(filename: str, max_length: int = 200) -> str:
    """Remove invalid characters from a string to make it a valid filename."""
    # Replace all invalid filename characters with underscores
    clean_name = INVALID_FILENAME_CHARS.sub("_", filename)
    # Collapse multiple consecutive underscores into one and strip leading/trailing underscores
    clean_name = re.sub(r"_+", "_", clean_name.strip()).strip("_")
    # Truncate the filename if it exceeds the maximum allowed length
    if len(clean_name) > max_length:
        clean_name = clean_name[:max_length]
    # Return the cleaned name, or a default fallback if the result is empty
    return clean_name or "unnamed_sheet"


def clean_cell_value(value, preserve_newlines: bool = False) -> str:
    """Convert any cell value to a clean string representation."""
    # Check if the value is NaN/missing using pandas, return empty string if so
    if pd.isna(value):
        return ""
    # Check if the value has a strftime method (likely a datetime object)
    if hasattr(value, "strftime"):
        try:
            # Format datetime objects into a standard readable date-time string
            return value.strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, TypeError):
            # Fall back to simple string conversion if formatting fails
            return str(value)

    # Convert the value to a string and strip leading/trailing whitespace
    str_val = str(value).strip()
    # Normalize Windows and old Mac line endings to Unix-style newlines
    str_val = str_val.replace("\r\n", "\n").replace("\r", "\n")
    # Replace tab characters with single spaces for consistent spacing
    str_val = str_val.replace("\t", " ")
    # If newlines should not be preserved, collapse all whitespace into single spaces
    if not preserve_newlines:
        str_val = " ".join(str_val.split())
    # Return the fully cleaned string value
    return str_val


def is_non_empty(value) -> bool:
    """Return True when a value should be treated as visible content."""
    # Return True only if the value is not NaN AND contains non-whitespace characters
    return pd.notna(value) and str(value).strip() != ""


def is_meaningful_row(row: pd.Series, threshold: int = 1) -> bool:
    """Check if a row contains meaningful data."""
    # Count non-empty cells in the row and compare against the minimum threshold
    return sum(1 for val in row if is_non_empty(val)) >= threshold


def make_unique_headers(headers: List[str]) -> List[str]:
    """Create non-empty, unique column names."""
    # Dictionary to track how many times each header name has been seen
    seen: Dict[str, int] = {}
    # List to store the final unique header names
    unique_headers = []

    # Iterate through each header name provided
    for name in headers:
        # If header is empty/None, generate a default name based on column position
        clean_name = name or f"Column_{len(unique_headers) + 1}"
        # Check if this header name has already been used
        if clean_name in seen:
            # Increment the counter for this duplicate name
            seen[clean_name] += 1
            # Append a numeric suffix to make the name unique
            clean_name = f"{clean_name}_{seen[clean_name]}"
        else:
            # Mark this name as seen for the first time
            seen[clean_name] = 1
        # Add the cleaned (and potentially renamed) header to the output list
        unique_headers.append(clean_name)

    # Return the list of unique, non-empty header names
    return unique_headers


def get_non_empty_blocks(df: pd.DataFrame) -> List[pd.DataFrame]:
    """Split a messy sheet into blocks separated by fully empty rows."""
    # Remove columns that are entirely empty to simplify processing
    df_cleaned = df.dropna(how="all", axis=1)
    # If no data remains after cleaning, return an empty list
    if df_cleaned.empty:
        return []

    # List to store the separated data blocks
    blocks = []
    # Temporary list to accumulate rows for the current block
    current_rows = []

    # Iterate through each row in the cleaned DataFrame
    for _, row in df_cleaned.iterrows():
        # If the row contains meaningful data, add it to the current block
        if is_meaningful_row(row):
            current_rows.append(row)
        # If the row is empty and we have accumulated rows, save the current block
        elif current_rows:
            # Convert accumulated rows to a DataFrame and remove empty columns
            blocks.append(pd.DataFrame(current_rows).dropna(how="all", axis=1))
            # Reset the accumulator for the next block
            current_rows = []

    # After the loop, save any remaining accumulated rows as a final block
    if current_rows:
        blocks.append(pd.DataFrame(current_rows).dropna(how="all", axis=1))

    # Return only non-empty blocks, filtering out any that ended up blank
    return [block for block in blocks if not block.empty]


def detect_header_row(block: pd.DataFrame) -> int:
    """Pick the row with the highest filled-cell count near the top of a block."""
    # Limit the search to the first 8 rows or the total rows if fewer than 8
    scan_limit = min(8, len(block))
    # List to store tuples of (row position, count of non-empty cells)
    scores = []
    # Iterate through each row position in the scan range
    for position in range(scan_limit):
        # Get the row data at the current position
        row = block.iloc[position]
        # Count non-empty cells in this row and store with its position
        scores.append((position, sum(1 for value in row if is_non_empty(value))))
    # Return the position of the row with the most non-empty cells (or 0 if no scores)
    return max(scores, key=lambda item: item[1])[0] if scores else 0


def extract_table_from_block(block: pd.DataFrame) -> pd.DataFrame:
    """Convert one non-empty block into a table-like DataFrame."""
    # If the block has no data, return an empty DataFrame
    if block.empty:
        return pd.DataFrame()

    # Detect which row is most likely the header row
    header_position = detect_header_row(block)
    # Extract all rows from the header position onwards to form the table
    table = block.iloc[header_position:].copy()
    # If no rows remain after extraction, return an empty DataFrame
    if table.empty:
        return pd.DataFrame()

    # Clean the values from the first row to use as column headers
    headers = [clean_cell_value(value) for value in table.iloc[0]]
    # Ensure all headers are unique and non-empty
    table.columns = make_unique_headers(headers)
    # Remove the header row from the data and reset the index
    table = table.iloc[1:].reset_index(drop=True)
    # Filter out rows that don't contain meaningful data
    table = table[table.apply(is_meaningful_row, axis=1)]
    # Remove any columns that are now entirely empty
    table = table.dropna(how="all", axis=1)

    # Handle edge case: if table is empty but the original block had one row
    if table.empty and len(block) == 1:
        # Use that single row as the table data
        row = block.iloc[[0]].copy().reset_index(drop=True)
        # Assign generic column names since no clear header existed
        row.columns = [f"Column_{i + 1}" for i in range(len(row.columns))]
        # Return the single-row table
        return row

    # Return the fully extracted and cleaned table
    return table


def extract_tabular_blocks(df: pd.DataFrame) -> List[pd.DataFrame]:
    """Extract all meaningful table/content blocks from a messy DataFrame."""
    # List to store all extracted tables
    tables = []
    # Process each non-empty block found in the DataFrame
    for block in get_non_empty_blocks(df):
        # Extract a structured table from the block
        table = extract_table_from_block(block)
        # If a valid table was extracted, add it to the results
        if not table.empty:
            tables.append(table)
    # Return all successfully extracted tables
    return tables


def wrap_text(value: str, width: int) -> List[str]:
    """Wrap a cell value to fit a fixed-width column."""
    # Clean the cell value while preserving intentional newline characters
    value = clean_cell_value(value, preserve_newlines=True)
    # If the value is empty after cleaning, return a single empty line
    if value == "":
        return [""]

    # List to store the wrapped lines of text
    wrapped_lines = []
    # Split the value by newlines to handle each paragraph separately
    for paragraph in value.split("\n"):
        # If a paragraph is empty (consecutive newlines), preserve it as an empty line
        if paragraph == "":
            wrapped_lines.append("")
            continue
        # Use textwrap to break the paragraph into lines of the specified width
        wrapped_lines.extend(
            textwrap.wrap(
                paragraph,
                # Ensure width is at least 1 to avoid errors
                width=max(width, 1),
                # Allow breaking long words that exceed the column width
                break_long_words=True,
                # Prevent breaking words at hyphens to maintain readability
                break_on_hyphens=False,
            )
            # If textwrap returns empty (shouldn't happen), provide a fallback
            or [""]
        )
    # Return the list of wrapped text lines
    return wrapped_lines


def render_wrapped_row(values: List[str], widths: List[int], separator: str = "  ") -> List[str]:
    """Render one logical row as one or more physical text lines."""
    # Wrap each cell's text according to its assigned column width
    wrapped_cells = [wrap_text(value, width) for value, width in zip(values, widths)]
    # Determine the height of the rendered row (max lines among all cells)
    row_height = max(len(cell_lines) for cell_lines in wrapped_cells)
    # List to store the final rendered text lines
    lines = []

    # Iterate through each line index up to the row height
    for line_index in range(row_height):
        # List to store the text segment from each cell at this line index
        parts = []
        # Iterate through each cell's wrapped lines and its target width
        for cell_lines, width in zip(wrapped_cells, widths):
            # Get the text for this line index, or empty string if cell has fewer lines
            text = cell_lines[line_index] if line_index < len(cell_lines) else ""
            # Left-align the text within its column width and add to parts
            parts.append(f"{text:<{width}}")
        # Join all cell segments with the separator and remove trailing whitespace
        lines.append(separator.join(parts).rstrip())

    # Return all rendered lines that make up this row
    return lines


def format_fixed_width(df: pd.DataFrame, max_column_width: int = DEFAULT_MAX_COLUMN_WIDTH) -> str:
    """Format a DataFrame as a wrapped fixed-width table."""
    # If the DataFrame has no data, return an empty string
    if df.empty:
        return ""

    # Clean all cell values in the DataFrame while preserving newlines
    clean_df = df.map(lambda value: clean_cell_value(value, preserve_newlines=True))
    # Ensure column names are unique and stripped of extra whitespace
    unique_col_names = make_unique_headers([str(col).strip() for col in clean_df.columns])
    # Assign the unique column names back to the DataFrame
    clean_df.columns = unique_col_names

    # List to store the calculated width for each column
    col_widths = []
    # Iterate through each column name and its index
    for i, column_name in enumerate(unique_col_names):
        # Get the column name and all cell values in this column as strings
        values = [column_name] + clean_df.iloc[:, i].astype(str).tolist()
        # Find the length of the longest single line in any value in this column
        longest_line = max(
            (len(line) for value in values for line in clean_cell_value(value, True).split("\n")),
            # Default to 0 if the column is completely empty
            default=0,
        )
        # Calculate width: clamp between MIN_COLUMN_WIDTH and max_column_width
        col_widths.append(min(max(longest_line, MIN_COLUMN_WIDTH), max_column_width))

    # List to store all lines of the formatted table output
    lines = []
    # Render the header row using the calculated column widths
    lines.extend(render_wrapped_row(unique_col_names, col_widths))
    # Add a separator line made of dashes matching each column's width
    lines.append("  ".join("-" * width for width in col_widths).rstrip())

    # Iterate through each data row in the cleaned DataFrame
    for _, row in clean_df.iterrows():
        # Extract each cell value from the row as a string
        values = [str(row.iloc[i]) for i in range(len(unique_col_names))]
        # Render the row (potentially multiple lines) and add to output
        lines.extend(render_wrapped_row(values, col_widths))

    # Join all lines with newlines to create the final formatted table string
    return "\n".join(lines)


def format_tables(df_raw: pd.DataFrame, sheet_name: str, max_column_width: int) -> str:
    """Format all detected content blocks in table mode."""
    # Extract all tabular data blocks from the raw DataFrame
    tables = extract_tabular_blocks(df_raw)
    # If no tables were detected, fall back to raw export
    if not tables:
        logger.info("  No clear table detected. Using raw export.")
        # Remove fully empty rows and columns for cleaner raw output
        raw = df_raw.dropna(how="all").dropna(how="all", axis=1)
        # Return the raw data as a string with a descriptive header
        return f"Raw content for sheet: {sheet_name}\n\n{raw.to_string(index=False)}"

    # Log how many content blocks were successfully detected
    logger.info(f"  Detected {len(tables)} content block(s)")
    # List to store formatted sections for each table
    sections = []
    # Iterate through each extracted table with a 1-based index
    for index, table in enumerate(tables, start=1):
        # Log the dimensions of the current table block
        logger.info(f"  Block {index}: {len(table)} rows x {len(table.columns)} columns")
        # Add a block title only if there are multiple tables
        title = f"BLOCK {index}" if len(tables) > 1 else ""
        # Format the table using fixed-width column alignment
        body = format_fixed_width(table, max_column_width=max_column_width)
        # Combine title and body, removing any extra whitespace, and add to sections
        sections.append(f"{title}\n{body}".strip())

    # Join all sections with double newlines for separation
    return "\n\n".join(sections)


def get_covered_merged_cells(worksheet) -> Dict[Tuple[int, int], bool]:
    """Identify merged cells whose value is visually owned by the top-left cell."""
    # Dictionary to track cells that are covered by a merge (non-top-left cells)
    covered_cells = {}
    # Iterate through each merged cell range in the worksheet
    for merged_range in worksheet.merged_cells.ranges:
        # Iterate through each row in the merged range
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            # Iterate through each column in the merged range
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                # Skip the top-left cell (it holds the actual value)
                if row != merged_range.min_row or col != merged_range.min_col:
                    # Mark this cell as covered by the merge
                    covered_cells[(row, col)] = True
    # Return the dictionary of covered cell coordinates
    return covered_cells


def get_layout_cell_value(worksheet, row: int, col: int, covered_cells: Dict[Tuple[int, int], bool]) -> str:
    """Return the visible text for one layout cell."""
    # If this cell is covered by a merged cell, return empty string (no visible content)
    if covered_cells.get((row, col)):
        return ""
    # Get the cell's value, clean it, and return the text representation
    return clean_cell_value(worksheet.cell(row, col).value, preserve_newlines=True)


def get_used_bounds(worksheet, covered_cells: Dict[Tuple[int, int], bool]):
    """Find the smallest visible grid rectangle containing data."""
    # Lists to track row and column indices that contain visible data
    rows = []
    cols = []
    # Iterate through every row in the worksheet
    for row in range(1, worksheet.max_row + 1):
        # Iterate through every column in the worksheet
        for col in range(1, worksheet.max_column + 1):
            # Get the visible text for this cell (respecting merged cells)
            value = get_layout_cell_value(worksheet, row, col, covered_cells)
            # If the cell has visible content, record its row and column
            if value:
                rows.append(row)
                cols.append(col)

    # If no data was found in any cell, return None
    if not rows or not cols:
        return None

    # Return the bounding box as (min_row, max_row, min_col, max_col)
    return min(rows), max(rows), min(cols), max(cols)


def excel_column_width_to_chars(width) -> int:
    """Convert an Excel column width into a practical text width."""
    # If Excel reports no specific width, return a reasonable default
    if width is None:
        return 12
    # Round the Excel width to the nearest integer and ensure minimum width
    return max(MIN_COLUMN_WIDTH, int(round(float(width))))


def format_layout_sheet(worksheet, max_column_width: int) -> str:
    """Approximate the visible Excel grid using worksheet values and column widths."""
    # Identify all cells covered by merged ranges
    covered_cells = get_covered_merged_cells(worksheet)
    # Find the bounding box of cells that contain visible data
    bounds = get_used_bounds(worksheet, covered_cells)
    # If no data exists in the worksheet, return an empty string
    if bounds is None:
        return ""

    # Unpack the bounding box coordinates
    min_row, max_row, min_col, max_col = bounds
    # List to store the calculated width for each column in the layout
    widths = []
    # Iterate through each column in the bounding box
    for col in range(min_col, max_col + 1):
        # Get the Excel column letter (e.g., 1 -> 'A')
        letter = get_column_letter(col)
        # Get the column width as defined in Excel
        excel_width = worksheet.column_dimensions[letter].width
        # Track the maximum observed content width in this column
        observed_width = 0
        # Iterate through each row in the bounding box
        for row in range(min_row, max_row + 1):
            # Get the visible text for this cell
            value = get_layout_cell_value(worksheet, row, col, covered_cells)
            # Update observed_width to the longest line in any cell in this column
            observed_width = max(
                observed_width,
                # Find the longest line in the cell value (split by newlines)
                max((len(line) for line in value.split("\n")), default=0),
            )
        # Use the larger of Excel's width or observed content, capped at max_column_width
        width = max(excel_column_width_to_chars(excel_width), min(observed_width, max_column_width))
        # Ensure the width doesn't exceed the global maximum
        widths.append(min(width, max_column_width))

    # List to store all rendered lines of the layout
    lines = []
    # Iterate through each row in the bounding box
    for row in range(min_row, max_row + 1):
        # List to store cell values for this row
        values = []
        # Iterate through each column in the bounding box
        for col in range(min_col, max_col + 1):
            # Get the visible text for this cell
            value = get_layout_cell_value(worksheet, row, col, covered_cells)
            # Add the cell value to the row values list
            values.append(value)
        # Render the row with proper wrapping and column alignment
        rendered = render_wrapped_row(values, widths)
        # Add all rendered lines to the output
        lines.extend(rendered)

    # Join all lines with newlines, strip trailing whitespace from each line and the result
    return "\n".join(line.rstrip() for line in lines).rstrip()


def get_output_file(directory: Path, base_name: str, extension: str, overwrite: bool) -> Path:
    """Return the output file path, optionally avoiding existing files."""
    # Construct the initial output file path
    output_file = directory / f"{base_name}.{extension}"
    # If overwrite is enabled, return this path directly
    if overwrite:
        return output_file

    # Counter for generating numbered filenames when avoiding overwrites
    counter = 1
    # While a file already exists at this path, generate a new path with a counter
    while output_file.exists():
        output_file = directory / f"{base_name}_{counter}.{extension}"
        counter += 1
    # Return the first available non-existing file path
    return output_file


def setup_pdf_font(pdf: FPDF):
    """Try to load a Unicode-capable font, fallback to helvetica."""
    # List of common Unicode font file paths across different operating systems
    font_paths = [
        "C:\\Windows\\Fonts\\arial.ttf",  # Windows
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",  # Linux
        "/System/Library/Fonts/Supplemental/Arial.ttf"  # macOS
    ]
    
    # Flag to track if a Unicode font was successfully loaded
    font_loaded = False
    # Try each font path in order
    for path in font_paths:
        # Check if the font file exists at this path
        if Path(path).exists():
            try:
                # Attempt to add the font to the PDF with a custom name
                pdf.add_font("ArialUnicode", "", path)
                # Set the PDF to use this font at size 8
                pdf.set_font("ArialUnicode", "", 8)
                # Mark the font as successfully loaded
                font_loaded = True
                # Stop trying other fonts
                break
            except:
                # If loading fails, continue to the next font path
                continue
    
    # If no Unicode font was loaded, fall back to the built-in helvetica font
    if not font_loaded:
        pdf.set_font("helvetica", "", 8)


def export_to_pdf(worksheet, output_path: Path, title: str):
    """Generate a high-fidelity PDF of an Excel sheet with borders and alignment."""
    # Create a new PDF in landscape orientation, using millimeters, A4 paper size
    pdf = FPDF(orientation="landscape", unit="mm", format="A4")
    # Enable automatic page breaks with a 15mm bottom margin
    pdf.set_auto_page_break(auto=True, margin=15)
    # Add the first page to the PDF
    pdf.add_page()
    
    # Set font to bold helvetica at size 14 for the title
    pdf.set_font("helvetica", "B", 14)
    # Add the title text centered, moving to the next line afterward
    pdf.cell(0, 10, title, ln=True, align="C")
    # Add 5mm of vertical spacing after the title
    pdf.ln(5)

    # Identify cells covered by merged ranges in the worksheet
    covered_cells = get_covered_merged_cells(worksheet)
    # Find the bounding box of visible data in the worksheet
    bounds = get_used_bounds(worksheet, covered_cells)
    # If no data exists, exit the function early
    if bounds is None:
        return

    # Unpack the bounding box coordinates
    min_row, max_row, min_col, max_col = bounds
    
    # List to store the calculated width for each PDF column
    col_widths = []
    # Iterate through each column in the bounding box
    for col in range(min_col, max_col + 1):
        # Get the Excel column letter
        letter = get_column_letter(col)
        # Get the column width from Excel (default to 12 if not set)
        excel_width = worksheet.column_dimensions[letter].width or 12
        # Scale the Excel width for PDF (multiply by 2, minimum 10mm)
        col_widths.append(max(10, excel_width * 2))

    # Calculate the total width of all columns combined
    total_width = sum(col_widths)
    # Calculate the available width on the PDF page (page width minus margins)
    page_width = pdf.w - 2 * pdf.l_margin
    # If columns exceed the page width, scale them down proportionally
    if total_width > page_width:
        # Calculate the scaling factor
        scale = page_width / total_width
        # Apply the scale factor to each column width
        col_widths = [w * scale for w in col_widths]

    # Set up the PDF font (try Unicode, fallback to helvetica)
    setup_pdf_font(pdf)
    
    # Iterate through each row in the bounding box
    for row in range(min_row, max_row + 1):
        # Track the maximum number of text lines needed for any cell in this row
        max_lines = 1
        # List to store cell values for this row
        row_values = []
        # Iterate through each column in the bounding box with its index
        for col_idx, col in enumerate(range(min_col, max_col + 1)):
            # Get the visible text for this cell
            val = get_layout_cell_value(worksheet, row, col, covered_cells)
            # Add the value to the row values list
            row_values.append(val)
            # If using the basic helvetica font, clean non-Latin characters
            if pdf.font_family == "helvetica":
                # Replace non-Latin-1 characters with placeholders, then decode back
                val = val.encode("latin-1", "replace").decode("latin-1")
            
            # Use multi_cell to calculate how many lines this text will occupy
            lines = len(pdf.multi_cell(col_widths[col_idx], 5, val, split_only=True))
            # Update max_lines if this cell requires more lines
            max_lines = max(max_lines, lines)
        
        # Calculate the total height needed for this row
        row_height = max_lines * 5
        # If the row would exceed the page, add a new page
        if pdf.get_y() + row_height > pdf.page_break_trigger:
            pdf.add_page()

        # Save the current X and Y coordinates (start of the row)
        start_x = pdf.get_x()
        start_y = pdf.get_y()
        
        # Iterate through each cell value in the row with its column index
        for col_idx, val in enumerate(row_values):
            # Save the current cursor position
            x = pdf.get_x()
            y = pdf.get_y()
            
            # If this cell is covered by a merged cell, just draw the border
            if covered_cells.get((row, min_col + col_idx)):
                # Draw a rectangle to represent the empty merged cell
                pdf.rect(x, y, col_widths[col_idx], row_height)
                # Move the cursor to the right edge of this cell
                pdf.set_x(x + col_widths[col_idx])
                # Skip to the next cell
                continue

            # Clean the value for the current font (if using helvetica)
            if pdf.font_family == "helvetica":
                val = val.encode("latin-1", "replace").decode("latin-1")

            # Draw the cell content with a border, left-aligned
            pdf.multi_cell(col_widths[col_idx], 5, val, border=1, align='L')
            # Move the cursor to the right edge of this cell at the original Y position
            pdf.set_xy(x + col_widths[col_idx], y)
            
        # Move the cursor to the start of the next row
        pdf.set_xy(start_x, start_y + row_height)

    # Save the completed PDF to the specified output path
    pdf.output(str(output_path))


def write_text_file(path: Path, content: str) -> None:
    """Write content to disk."""
    # Write the content string to the file using UTF-8 encoding
    path.write_text(content, encoding="utf-8")


def process_csv(input_path: Path, output_folder: Path, max_column_width: int, overwrite: bool) -> bool:
    """Convert one CSV file into .txt and .pdf files."""
    # Create the output directory structure based on the input file's name
    file_output_root = output_folder / sanitize_filename(input_path.stem)
    # Define the subdirectory for TXT output files
    txt_dir = file_output_root / "TXT"
    # Define the subdirectory for PDF output files
    pdf_dir = file_output_root / "PDF"
    
    # Create the TXT directory (and parents if needed), silently skip if it exists
    txt_dir.mkdir(parents=True, exist_ok=True)
    # Create the PDF directory (and parents if needed), silently skip if it exists
    pdf_dir.mkdir(parents=True, exist_ok=True)

    # Log the CSV file being read
    logger.info(f"Reading CSV file: {input_path.name}")
    # Log where the output files will be saved
    logger.info(f"Output will be saved to: {file_output_root}")

    try:
        # Read the CSV file into a DataFrame without treating any row as a header
        df = pd.read_csv(input_path, header=None)
        # Check if the DataFrame is empty or contains only empty rows
        if df.empty or df.dropna(how="all").empty:
            logger.warning("CSV file is empty.")
            # Return False to indicate no data was processed
            return False

        # Format the data into table-style text content
        content = format_tables(df, input_path.stem, max_column_width)
        
        # Determine the output path for the TXT file
        txt_path = get_output_file(txt_dir, sanitize_filename(input_path.stem), "txt", overwrite)
        # Write the formatted content to the TXT file
        write_text_file(txt_path, content)
        # Log the successful TXT file save
        logger.info(f"  Saved TXT: {txt_path.name}")
        
        # Determine the output path for the PDF file
        pdf_path = get_output_file(pdf_dir, sanitize_filename(input_path.stem), "pdf", overwrite)
        # Create a new PDF in landscape orientation, millimeters, A4 size
        pdf = FPDF(orientation="landscape", unit="mm", format="A4")
        # Add the first page to the PDF
        pdf.add_page()
        # Set up the font (try Unicode, fallback to helvetica)
        setup_pdf_font(pdf)
        # Iterate through each line of the formatted text content
        for line in content.split('\n'):
            # If using helvetica font, clean non-Latin characters
            if pdf.font_family == "helvetica":
                line = line.encode("latin-1", "replace").decode("latin-1")
            # Add the line as a cell in the PDF
            pdf.cell(0, 5, line, ln=True)
        # Save the completed PDF to the output path
        pdf.output(str(pdf_path))
        # Log the successful PDF file save
        logger.info(f"  Saved PDF: {pdf_path.name}")
        
        # Return True to indicate successful processing
        return True
    except Exception as exc:
        # Log any errors that occurred during CSV processing
        logger.error(f"Failed to process CSV file: {exc}")
        # Return False to indicate processing failure
        return False


def process_excel(
    input_path: Path,
    output_folder: Path,
    mode: str,
    max_column_width: int,
    overwrite: bool,
) -> bool:
    """Read an Excel workbook and convert each sheet to .txt and .pdf."""
    # Create the output directory structure based on the input file's name
    file_output_root = output_folder / sanitize_filename(input_path.stem)
    # Define the subdirectory for TXT output files
    txt_dir = file_output_root / "TXT"
    # Define the subdirectory for PDF output files
    pdf_dir = file_output_root / "PDF"
    
    # Create the TXT directory (and parents if needed), silently skip if it exists
    txt_dir.mkdir(parents=True, exist_ok=True)
    # Create the PDF directory (and parents if needed), silently skip if it exists
    pdf_dir.mkdir(parents=True, exist_ok=True)

    # Log the Excel file being read
    logger.info(f"Reading Excel file: {input_path.name}")
    # Log where the output files will be saved
    logger.info(f"Output will be saved to: {file_output_root}")

    try:
        # If layout mode is requested but file is .xls (old format), fall back to table mode
        if mode == "layout" and input_path.suffix.lower() == ".xls":
            logger.warning("Layout mode requires openpyxl-supported workbooks. Falling back to table mode for .xls.")
            mode = "table"

        # Load the Excel file with pandas only if table mode is active
        excel_data = pd.ExcelFile(input_path) if mode == "table" else None
        # Load the workbook with openpyxl to access raw sheet data and formatting
        workbook = load_workbook(input_path, data_only=True)
        # Get the list of all sheet names in the workbook
        sheet_names = workbook.sheetnames

        # If the workbook has no sheets, log a warning and return False
        if not sheet_names:
            logger.warning("No sheets found in Excel file.")
            return False

        # Log how many sheets were found
        logger.info(f"Found {len(sheet_names)} sheet(s)")
        # Counter for successfully processed sheets
        processed_count = 0
        # Counter for sheets that were skipped due to errors
        skipped_count = 0

        # Iterate through each sheet name in the workbook
        for sheet_name in sheet_names:
            # Log which sheet is being processed
            logger.info(f"\nProcessing sheet: '{sheet_name}'...")
            try:
                # 1. Generate Content for TXT
                # If layout mode, format the sheet preserving its visual layout
                if mode == "layout":
                    content = format_layout_sheet(workbook[sheet_name], max_column_width)
                # Otherwise, use table mode to extract data blocks
                else:
                    df_raw = pd.read_excel(input_path, sheet_name=sheet_name, header=None)
                    content = format_tables(df_raw, sheet_name, max_column_width)

                # If there is actual content to write
                if content.strip():
                    # Determine the output path for the TXT file
                    txt_path = get_output_file(txt_dir, sanitize_filename(sheet_name), "txt", overwrite)
                    # Write the content to the TXT file
                    write_text_file(txt_path, content)
                    # Log the successful TXT file save
                    logger.info(f"  Saved TXT: {txt_path.name}")
                
                # 2. Generate High-Fidelity PDF
                # Determine the output path for the PDF file
                pdf_path = get_output_file(pdf_dir, sanitize_filename(sheet_name), "pdf", overwrite)
                # Export the sheet to a formatted PDF
                export_to_pdf(workbook[sheet_name], pdf_path, sheet_name)
                # Log the successful PDF file save
                logger.info(f"  Saved PDF: {pdf_path.name}")

                # Increment the processed sheet counter
                processed_count += 1
            except Exception as exc:
                # Log any error that occurred while processing this sheet
                logger.error(f"  Error processing '{sheet_name}': {exc}")
                # Increment the skipped sheet counter
                skipped_count += 1

        # Print a separator line for cleaner log output
        logger.info(f"\n{'=' * 50}")
        # Log that processing is complete
        logger.info("Processing complete!")
        # Log how many sheets were successfully processed
        logger.info(f"  Processed: {processed_count} sheet(s)")
        # Log how many sheets were skipped
        logger.info(f"  Skipped: {skipped_count} sheet(s)")
        # Log the output directory location
        logger.info(f"  Location: {file_output_root}")
        # Print another separator line
        logger.info(f"{'=' * 50}")
        # Return True if at least one sheet was processed successfully
        return processed_count > 0
    except Exception as exc:
        # Log any fatal errors that occurred during Excel processing
        logger.error(f"Failed to process Excel file: {exc}")
        # Return False to indicate overall failure
        return False


def process_input(input_file: str, output_folder: str, mode: str, max_column_width: int, overwrite: bool) -> bool:
    """Dispatch processing based on file type."""
    # Convert the input file path string to a Path object
    input_path = Path(input_file)
    # Convert the output folder path string to a Path object
    output_path = Path(output_folder)

    # Check if the input file actually exists on disk
    if not input_path.exists():
        logger.error(f"File not found: '{input_file}'")
        # Return False since the file cannot be processed
        return False

    # Get the file extension in lowercase for case-insensitive comparison
    extension = input_path.suffix.lower()
    # Check if the file extension is in the list of supported types
    if extension not in SUPPORTED_INPUT_EXTENSIONS:
        logger.error(f"Unsupported file type: '{extension}'. Use Excel or CSV files.")
        # Return False since the file type is not supported
        return False

    # Create the output directory (and parents if needed), silently skip if it exists
    output_path.mkdir(parents=True, exist_ok=True)

    # If the file is a CSV, delegate to the CSV processing function
    if extension == ".csv":
        return process_csv(input_path, output_path, max_column_width, overwrite)

    # Otherwise, delegate to the Excel processing function
    return process_excel(input_path, output_path, mode, max_column_width, overwrite)


def main():
    # Create an argument parser with a description of the script
    parser = argparse.ArgumentParser(
        description="Convert Excel sheets or CSV files to formatted .txt files.",
        # Allow the description to preserve its formatting
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    # Add a required positional argument for the input file path
    parser.add_argument("input", help="Path to the source Excel or CSV file")
    # Add an optional argument for the output folder, defaulting to "data/Output"
    parser.add_argument("-o", "--output", default="data/Output", help="Base folder for output (default: data/Output)")
    # Add an optional argument to choose between "table" and "layout" export modes
    parser.add_argument(
        "--mode",
        choices=("table", "layout"),
        default="layout",
        help="Export style: table extracts data blocks; layout approximates the Excel grid (default: layout)",
    )
    # Add an optional argument to set the maximum column width before text wrapping
    parser.add_argument(
        "--max-column-width",
        type=int,
        default=DEFAULT_MAX_COLUMN_WIDTH,
        help=f"Maximum text width per column before wrapping (default: {DEFAULT_MAX_COLUMN_WIDTH})",
    )
    # Add a flag to prevent overwriting existing files (creates numbered copies instead)
    parser.add_argument(
        "--no-overwrite",
        action="store_true",
        help="Keep existing files and create numbered names instead of overwriting",
    )
    # Add a flag to enable verbose (debug-level) logging
    parser.add_argument("-v", "--verbose", action="store_true", help="Enable verbose logging")

    # Parse the command-line arguments provided by the user
    args = parser.parse_args()
    # If verbose mode is enabled, set the logging level to DEBUG for more detailed output
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Ensure the max column width is at least the defined minimum
    max_column_width = max(args.max_column_width, MIN_COLUMN_WIDTH)
    # Call the main processing function with all parsed arguments
    success = process_input(
        args.input,
        args.output,
        args.mode,
        max_column_width,
        # Invert no_overwrite flag: True means overwrite is allowed
        overwrite=not args.no_overwrite,
    )
    # Exit the program with code 0 on success, or code 1 on failure
    raise SystemExit(0 if success else 1)


# Standard Python idiom: execute main() only when the script is run directly
if __name__ == "__main__":
    main()
